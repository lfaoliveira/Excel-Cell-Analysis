from openpyxl import load_workbook




class Participante:
  def __init__(self, nome, tempo):
    self.nome = nome
    self.tempo = tempo



def calcula_tempo(string):
  ##EM SEGUNDOS ##
  tempo = 0
  string = string.split(' ')
  aux_min = [0,0]
  aux_sec = [0,0]
  
  for cont in range(len(string)):
    sub = string[cont]
    for i in range(len(sub)):
      if(sub == "min"):
        continue

      ###
      char = sub[i]
      
      if(char == 'h'):
        ## pega o numero antes de 'h' e converte
        hora = int(sub[i - 1])
        tempo += 3600*hora
        ## pega o numero antes de 'm' e converte
      elif(char == 'm'):
        aux_min[1] = int(sub[i - 1])
        try:
          aux_min[0] = int(sub[i - 2])
        except:
          aux_min[0] = 0
        ## pega o numero antes de 's' e converte
      elif(char == 's'):
        aux_sec[1] = int(sub[i - 1])
        try:
          aux_sec[0] = int(sub[i - 2])
        except:
          aux_sec[0] = 0

  ## CONVERSAO MATEMATICA ##
  d = 10
  for i in range(2):
    tempo += d*aux_min[i]  * 60
    d = d//10
  
  d = 10
  for i in range(2):
    tempo += d*aux_sec[i]
    d = d//10

  return tempo



  
## EM SEGUNDOS ##
def formata_tempo(tempo):
  tempo = int(tempo)
  hora = tempo // 3600
  tempo -= 3600*hora

  min = tempo//60
  tempo -= (min)*60


  sec = tempo
  hora = str(hora)
  min = str(min)
  sec = str(sec)

  string = hora + "h" + " " + min + "m " + sec +"s"
  
  return string


#########################
#########################
# OBS: NOMES PRECISAM ESTAR SEMPRE COM OS MESMOS CARACTERES, SEM ESPAÇO OU CARACTERES ADICIONAIS 
#############################


file = input("NOme do arquivo:")
pasta  = load_workbook(file)
nome_pasta = input("Nome da planilha:")

planilha1 = pasta[nome_pasta]
planilha1 = pasta.active




lista_part = list(())

x = 11
n = x + 1
## loop para ajeitar o x ##

while(True):
  celula  = planilha1.cell(row = n, column = 4)
  string = celula.value
  if(string == None or string == "NULL"):
    break
  n += 1
  
x = n + 3

## LOOP PARA CALCULOS ##

while(True):
  celula  = planilha1.cell(row = x, column = 4)
  string = celula.value
  
  if (string == None or string == "NULL"):
    break
  
  
  nome = planilha1.cell(row = x,column = 1).value
  nalista = False 
  calculado = False
  
## Procura na lista o nome e adiciona mais tempo se encontrar

  for elem in lista_part:
    
    if(elem.nome == nome):
      elem.tempo += calcula_tempo(string)
      nalista = True
      calculado = True
  if(nalista == False):
    part = Participante(nome, 0)
  


  ## calculo padrao ( se o part. nao estiver na lista) ##
  if calculado == False:
      tempo = calcula_tempo(string)
      part.tempo += tempo
      
  
  
    
  if(nalista == False):
    lista_part.append(part);
  
  
  x += 1

## fim while ##
  
  
  
pasta.create_sheet("Tempo")
planilha2 = pasta["Tempo"]

planilha2.cell(row = 4, column = 1, value = "Nome")
planilha2.cell(row = 4, column = 2, value = "Presença")

x = 5
for part in lista_part:
  part.tempo = formata_tempo(part.tempo)
  planilha2.cell(row = x, column = 1, value = part.nome)
  planilha2.cell(row = x, column = 2, value = part.tempo)
  
  x += 1
###


file = file.replace(".xlsx","")
file = file + ".xlsx"

pasta.save(filename = file)